"""
Parser principal d'extraits bancaires.
Lit CSV ou XLSX -> détecte la structure -> produit des transactions normalisées.
"""
import csv
import io
from .parsers.schema_detector import detect_schema, find_header_row
from .parsers.metadata_extractor import extract_metadata
from .normalize import parse_date, parse_amount, clean_description
from .classifier import classify


def _read_rows(content: bytes, filename: str) -> list[list]:
    """Retourne toutes les lignes brutes (liste de listes), CSV ou XLSX."""
    name = filename.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows
    else:
        # CSV / TSV
        text = content.decode("utf-8-sig")
        # détection naïve du séparateur
        delimiter = "\t" if text.count("\t") > text.count(",") else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        return [row for row in reader]


def _resolve_amount(row, mapping, mode) -> float | None:
    """Calcule le montant signé selon le mode détecté."""
    def cell(field):
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    if mode == "split":
        debit = parse_amount(cell("debit"))
        credit = parse_amount(cell("credit"))
        if credit:
            return abs(credit)
        if debit:
            return -abs(debit)
        return 0.0

    if mode == "typed":
        amt = parse_amount(cell("amount"))
        t = str(cell("type") or "").strip()
        if amt is None:
            return None
        # دائن = crédit (positif), مدين = débit (négatif)
        if "دائن" in t or "credit" in t.lower() or "cr" == t.lower():
            return abs(amt)
        if "مدين" in t or "debit" in t.lower() or "dr" == t.lower():
            return -abs(amt)
        return amt

    if mode == "signed":
        return parse_amount(cell("amount"))

    return None


def parse_statement(content: bytes, filename: str) -> dict:
    """
    Pipeline complet. Retourne :
      {
        detected: {bank_hint, mode, confidence, mapping},
        transactions: [...],
        summary: {total, classified, needs_review, inflow, outflow}
      }
    """
    raw_rows = _read_rows(content, filename)
    if not raw_rows:
        return {"error": "fichier vide", "transactions": []}

    header_idx = find_header_row(raw_rows)
    headers = [str(c) if c is not None else "" for c in raw_rows[header_idx]]
    schema = detect_schema(headers)
    mapping, mode = schema["mapping"], schema["mode"]

    # extraire les métadonnées des lignes AVANT le tableau
    metadata = extract_metadata(raw_rows[:header_idx], filename)

    transactions = []
    inflow = outflow = 0.0
    classified = needs_review = 0

    def cell(row, field):
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in raw_rows[header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        # ignorer lignes de pied de page (pas de date détectable)
        d = parse_date(row[mapping["date"]]) if "date" in mapping and mapping["date"] < len(row) else None
        if d is None:
            continue

        desc = clean_description(row[mapping["description"]]) if "description" in mapping and mapping["description"] < len(row) else ""
        amount = _resolve_amount(row, mapping, mode)
        if amount is None:
            continue

        # ignorer les lignes de solde initial : ce sont des soldes de référence,
        # pas des mouvements de trésorerie (montant toujours nul)
        OPENING_LABELS = ["رصيد افتتاحي", "opening balance", "solde initial",
                          "رصيد اول", "brought forward", "balance b/f"]
        if any(lbl.lower() in desc.lower() for lbl in OPENING_LABELS) and amount == 0:
            continue

        balance = None
        if "balance" in mapping and mapping["balance"] < len(row):
            balance = parse_amount(row[mapping["balance"]])

        # nouveaux champs exploités : référence et type de mouvement
        reference = clean_description(cell(row, "ref")) or None
        movement_type = clean_description(cell(row, "type")) or None

        cls = classify(desc, amount)
        if not cls["needs_review"]:
            classified += 1
        else:
            needs_review += 1

        if amount >= 0:
            inflow += amount
        else:
            outflow += abs(amount)

        transactions.append({
            "date": d.isoformat(),
            "description": desc,
            "amount": round(amount, 2),
            "balance": round(balance, 2) if balance is not None else None,
            "reference": reference,
            "movement_type": movement_type,
            "category": cls["category"],
            "counterparty": cls["counterparty"],
            "confidence": cls["confidence"],
            "needs_review": cls["needs_review"],
            "review_reason": cls.get("review_reason"),
        })

    return {
        "detected": {
            "mode": mode,
            "confidence": schema["confidence"],
            "mapping": mapping,
            "header_row": header_idx,
        },
        "metadata": metadata,
        "transactions": transactions,
        "summary": {
            "total": len(transactions),
            "classified": classified,
            "needs_review": needs_review,
            "inflow": round(inflow, 2),
            "outflow": round(outflow, 2),
            "net": round(inflow - outflow, 2),
        },
    }
